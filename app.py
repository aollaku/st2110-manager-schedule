
from flask import Flask, jsonify, render_template, request, send_file, redirect, url_for, session
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime
import shutil
import re
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
import threading
import time
import os
import subprocess
import platform
import ipaddress
import csv
from io import StringIO
import json
import tempfile

APP_DIR = Path(__file__).resolve().parent
DATA_DIR = APP_DIR / "data"
BACKUP_DIR = APP_DIR / "backups"
BACKUP_DIR.mkdir(exist_ok=True)


USERS_PATH = APP_DIR / "users.json"

ROLE_LEVELS = {
    "read": 1,
    "write": 2,
    "admin": 3
}


def init_users():
    if USERS_PATH.exists():
        return

    default_users = [
        {
            "username": "admin",
            "password_hash": generate_password_hash("admin123"),
            "role": "admin",
            "must_change_password": True,
            "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
    ]
    USERS_PATH.write_text(json.dumps(default_users, indent=2), encoding="utf-8")


def load_users():
    init_users()
    try:
        return json.loads(USERS_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []


def save_users(users):
    USERS_PATH.write_text(json.dumps(users, indent=2), encoding="utf-8")


def find_user(username):
    username = clean_text(username).lower()
    for user in load_users():
        if user.get("username", "").lower() == username:
            return user
    return None


def current_user():
    username = session.get("username")
    if not username:
        return None
    return find_user(username)


def has_role(required):
    user = current_user()
    if not user:
        return False
    return ROLE_LEVELS.get(user.get("role", "read"), 0) >= ROLE_LEVELS.get(required, 99)


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not current_user():
            if request.path.startswith("/api/"):
                return jsonify({"error": "Authentication required"}), 401
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper


def role_required(role):
    def decorator(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not current_user():
                return jsonify({"error": "Authentication required"}), 401
            if not has_role(role):
                return jsonify({"error": f"{role} permission required"}), 403
            return fn(*args, **kwargs)
        return wrapper
    return decorator


def can_write():
    return has_role("write")


def can_admin():
    return has_role("admin")




WORKBOOK_PATH = DATA_DIR / "Move_To_IP_Schedule_v5.96.xlsx"

app = Flask(__name__)
app.secret_key = os.environ.get('ST2110_SECRET_KEY', 'change-this-secret-key-for-production')

CACHE_LOCK = threading.Lock()
CACHE_READY = False
CACHE_BUILDING = False
CACHE_ERROR = ""
CACHE_RECORDS = []
CACHE_SHEETS = []
CACHE_BUILT_AT = ""

MAX_RECORDS_DEFAULT = 300


def clean_text(value):
    if value is None:
        return ""
    text = str(value).replace("\n", " ").replace("\r", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalise_header(text):
    text = clean_text(text)
    if not text:
        return ""
    if text.lower().startswith("unnamed"):
        return ""
    return text


def merged_value(ws, row, col):
    value = ws.cell(row, col).value
    if value is not None:
        return clean_text(value)

    merged_ranges = getattr(getattr(ws, "merged_cells", None), "ranges", [])
    for merged in merged_ranges:
        if merged.min_row <= row <= merged.max_row and merged.min_col <= col <= merged.max_col:
            return clean_text(ws.cell(merged.min_row, merged.min_col).value)
    return ""


def make_unique_headers(headers):
    seen = {}
    output = []
    for idx, h in enumerate(headers, start=1):
        h = normalise_header(h)
        if not h:
            h = f"Column {idx}"
        base = h
        if base in seen:
            seen[base] += 1
            h = f"{base} {seen[base]}"
        else:
            seen[base] = 1
        output.append(h)
    return output


def detect_header_rows(ws):
    best_row = 3
    best_score = -1

    for r in range(1, min(10, ws.max_row) + 1):
        values = [clean_text(ws.cell(r, c).value).lower() for c in range(1, min(ws.max_column, 80) + 1)]
        score = 0
        keywords = [
            "prefix", "no.", "no", "area", "loc.", "loc", "unit",
            "conn", "type", "function", "pins", "remarks", "ip", "vlan"
        ]
        for v in values:
            if v in keywords:
                score += 4
            elif v:
                score += 1
        if score > best_score:
            best_score = score
            best_row = r

    return max(1, best_row - 1), best_row


def get_clean_headers(ws):
    group_row, header_row = detect_header_rows(ws)
    headers = []
    current_group = ""

    for col in range(1, ws.max_column + 1):
        group = normalise_header(merged_value(ws, group_row, col))
        field = normalise_header(merged_value(ws, header_row, col))

        if group and len(group) < 80:
            current_group = group

        group_to_use = current_group
        if group_to_use.lower() in ["f", "from:", "to:", "moved from:", "moved to:"]:
            group_to_use = current_group

        if field and group_to_use and field.lower() not in group_to_use.lower():
            headers.append(f"{group_to_use} {field}")
        elif field:
            headers.append(field)
        elif group_to_use:
            headers.append(group_to_use)
        else:
            headers.append("")

    return make_unique_headers(headers), header_row


def is_data_row(values):
    filled = [clean_text(v) for v in values if clean_text(v)]
    if len(filled) < 2:
        return False

    joined = " ".join(filled).upper()
    skip = [
        "CONNECTION BETWEEN",
        "IP ROUTER",
        "GV IP ROUTER",
        "SHEET",
        "NOTES"
    ]
    if any(joined.startswith(s) for s in skip):
        return False

    return True


def compact_field_name(key):
    key = key.replace("Moved from: ", "From ")
    key = key.replace("Moved to: ", "To ")
    key = re.sub(r"\s+", " ", key).strip()
    return key


def row_to_record(sheet_name, ws, row_number, headers):
    raw_values = [ws.cell(row_number, col).value for col in range(1, ws.max_column + 1)]
    if not is_data_row(raw_values):
        return None

    fields = {}
    search_parts = [sheet_name]

    for col, header in enumerate(headers, start=1):
        value = clean_text(ws.cell(row_number, col).value)
        if value:
            clean_header = compact_field_name(header)
            fields[clean_header] = value
            search_parts.append(clean_header)
            search_parts.append(value)

    if not fields:
        return None

    prefix = fields.get("Prefix", "")
    no = fields.get("No.", fields.get("No", ""))
    cable_id = " ".join([prefix, no]).strip()
    if not cable_id:
        cable_id = fields.get("Cable", "") or fields.get("Cable ID", "") or f"Row {row_number}"

    source_parts = []
    dest_parts = []
    network_parts = []
    useful_parts = []

    for key, value in fields.items():
        lk = key.lower()
        short = key
        short = short.replace("From ", "").replace("To ", "")
        short = short.replace("Moved from: ", "").replace("Moved to: ", "")

        item = f"{short}: {value}"

        if "from" in lk or "moved from" in lk:
            source_parts.append(item)
        elif "to" in lk or "moved to" in lk:
            dest_parts.append(item)
        elif any(x in lk for x in ["ip", "vlan", "multicast", "sdp", "network"]):
            network_parts.append(item)
        elif lk not in ["prefix", "no.", "no"]:
            useful_parts.append(item)

    return {
        "sheet": sheet_name,
        "row_number": row_number,
        "cable_id": cable_id,
        "source": " | ".join(source_parts[:8]),
        "destination": " | ".join(dest_parts[:8]),
        "network": " | ".join(network_parts[:6]),
        "summary": " | ".join(useful_parts[:8]),
        "fields": fields,
        "headers": list(fields.keys()),
        "search": " ".join(search_parts).lower()
    }


def build_cache():
    global CACHE_READY, CACHE_BUILDING, CACHE_ERROR, CACHE_RECORDS, CACHE_SHEETS, CACHE_BUILT_AT

    with CACHE_LOCK:
        CACHE_BUILDING = True
        CACHE_READY = False
        CACHE_ERROR = ""

    try:
        start = time.time()

        wb = load_workbook(WORKBOOK_PATH, data_only=True)
        all_records = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            headers, header_row = get_clean_headers(ws)

            for row_number in range(header_row + 1, ws.max_row + 1):
                record = row_to_record(sheet_name, ws, row_number, headers)
                if record:
                    all_records.append(record)

        with CACHE_LOCK:
            CACHE_RECORDS = all_records
            CACHE_SHEETS = wb.sheetnames
            CACHE_READY = True
            CACHE_BUILDING = False
            CACHE_BUILT_AT = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            CACHE_ERROR = ""

        print(f"Cache built: {len(all_records)} records in {time.time() - start:.2f}s")

    except Exception as exc:
        with CACHE_LOCK:
            CACHE_READY = False
            CACHE_BUILDING = False
            CACHE_ERROR = str(exc)
        print(f"Cache build failed: {exc}")


def ensure_cache_thread():
    global CACHE_BUILDING
    with CACHE_LOCK:
        if CACHE_READY or CACHE_BUILDING:
            return
        CACHE_BUILDING = True

    thread = threading.Thread(target=build_cache, daemon=True)
    thread.start()


def backup_workbook():
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup = BACKUP_DIR / f"Move_To_IP_Schedule_backup_{stamp}.xlsx"
    shutil.copy2(WORKBOOK_PATH, backup)
    return backup




@app.route("/login", methods=["GET", "POST"])
def login():
    init_users()
    error = ""
    if request.method == "POST":
        username = clean_text(request.form.get("username", ""))
        password = request.form.get("password", "")
        user = find_user(username)
        if user and check_password_hash(user.get("password_hash", ""), password):
            session["username"] = user["username"]
            append_audit("login", {"username": user["username"]}) if "append_audit" in globals() else None
            return redirect(url_for("index"))
        error = "Invalid username or password"
    return render_template("login.html", error=error)


@app.route("/logout")
def logout():
    username = session.get("username")
    session.clear()
    if username and "append_audit" in globals():
        append_audit("logout", {"username": username})
    return redirect(url_for("login"))


@app.route("/api/me")
@login_required
def api_me():
    user = current_user()
    return jsonify({
        "username": user.get("username"),
        "role": user.get("role", "read"),
        "must_change_password": bool(user.get("must_change_password", False)),
        "can_write": can_write(),
        "can_admin": can_admin()
    })


@app.route("/api/change-password", methods=["POST"])
@login_required
def api_change_password():
    payload = request.get_json(force=True)
    old_password = payload.get("old_password", "")
    new_password = payload.get("new_password", "")

    if len(new_password) < 6:
        return jsonify({"error": "New password must be at least 6 characters"}), 400

    users = load_users()
    username = session.get("username")
    for user in users:
        if user.get("username") == username:
            if not check_password_hash(user.get("password_hash", ""), old_password):
                return jsonify({"error": "Old password is incorrect"}), 400
            user["password_hash"] = generate_password_hash(new_password)
            user["must_change_password"] = False
            save_users(users)
            append_audit("password_changed", {"username": username}) if "append_audit" in globals() else None
            return jsonify({"status": "changed"})
    return jsonify({"error": "User not found"}), 404


@app.route("/api/users")
@role_required("admin")
def api_users():
    users = load_users()
    safe = []
    for user in users:
        safe.append({
            "username": user.get("username"),
            "role": user.get("role", "read"),
            "must_change_password": bool(user.get("must_change_password", False)),
            "created_at": user.get("created_at", "")
        })
    return jsonify({"users": safe})


@app.route("/api/users", methods=["POST"])
@role_required("admin")
def api_create_user():
    payload = request.get_json(force=True)
    username = clean_text(payload.get("username", "")).lower()
    password = payload.get("password", "")
    role = payload.get("role", "read")

    if not username:
        return jsonify({"error": "Username is required"}), 400
    if len(password) < 6:
        return jsonify({"error": "Password must be at least 6 characters"}), 400
    if role not in ROLE_LEVELS:
        return jsonify({"error": "Invalid role"}), 400
    if find_user(username):
        return jsonify({"error": "User already exists"}), 400

    users = load_users()
    users.append({
        "username": username,
        "password_hash": generate_password_hash(password),
        "role": role,
        "must_change_password": True,
        "created_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    })
    save_users(users)
    append_audit("user_created", {"username": username, "role": role}) if "append_audit" in globals() else None
    return jsonify({"status": "created"})


@app.route("/api/users/<username>", methods=["DELETE"])
@role_required("admin")
def api_delete_user(username):
    username = clean_text(username).lower()
    if username == session.get("username", "").lower():
        return jsonify({"error": "You cannot delete your own logged-in account"}), 400
    users = load_users()
    before = len(users)
    users = [u for u in users if u.get("username", "").lower() != username]
    if len(users) == before:
        return jsonify({"error": "User not found"}), 404
    save_users(users)
    append_audit("user_deleted", {"username": username}) if "append_audit" in globals() else None
    return jsonify({"status": "deleted"})


@app.route("/api/users/<username>/role", methods=["POST"])
@role_required("admin")
def api_update_user_role(username):
    payload = request.get_json(force=True)
    role = payload.get("role", "read")
    if role not in ROLE_LEVELS:
        return jsonify({"error": "Invalid role"}), 400

    users = load_users()
    for user in users:
        if user.get("username", "").lower() == username.lower():
            user["role"] = role
            save_users(users)
            append_audit("user_role_changed", {"username": username, "role": role}) if "append_audit" in globals() else None
            return jsonify({"status": "updated"})
    return jsonify({"error": "User not found"}), 404


@app.route("/")
@login_required
def index():
    ensure_cache_thread()
    return render_template("index.html")


@app.route("/api/status")
@login_required
def api_status():
    ensure_cache_thread()
    with CACHE_LOCK:
        return jsonify({
            "ready": CACHE_READY,
            "building": CACHE_BUILDING,
            "error": CACHE_ERROR,
            "records": len(CACHE_RECORDS),
            "sheets": CACHE_SHEETS,
            "built_at": CACHE_BUILT_AT,
            "workbook": WORKBOOK_PATH.name,
            "port": 5005
        })


@app.route("/api/reload", methods=["POST"])
@login_required
def api_reload():
    thread = threading.Thread(target=build_cache, daemon=True)
    thread.start()
    return jsonify({"status": "reloading"})


@app.route("/api/sheets")
@login_required
def api_sheets():
    ensure_cache_thread()
    with CACHE_LOCK:
        return jsonify({"sheets": CACHE_SHEETS})


@app.route("/api/records")
@login_required
def api_records():
    ensure_cache_thread()

    with CACHE_LOCK:
        if not CACHE_READY:
            return jsonify({
                "ready": False,
                "building": CACHE_BUILDING,
                "error": CACHE_ERROR,
                "records": [],
                "count": 0
            }), 202

        query = clean_text(request.args.get("q", "")).lower()
        sheet = request.args.get("sheet", "ALL")
        limit = int(request.args.get("limit", MAX_RECORDS_DEFAULT))

        terms = [t for t in re.split(r"\s+", query) if t]
        output = []

        for record in CACHE_RECORDS:
            if sheet != "ALL" and record["sheet"] != sheet:
                continue

            haystack = record.get("search", "")
            if terms and not all(term in haystack for term in terms):
                continue

            visible = dict(record)
            visible.pop("search", None)
            output.append(visible)

            if limit and len(output) >= limit:
                break

        return jsonify({
            "ready": True,
            "records": output,
            "count": len(output),
            "query": query,
            "sheet": sheet,
            "limited": bool(limit and len(output) >= limit)
        })


@app.route("/api/record/<sheet>/<int:row_number>", methods=["POST"])
@role_required("write")
def save_record(sheet, row_number):
    payload = request.get_json(force=True)
    fields = payload.get("fields", {})

    backup = backup_workbook()

    wb = load_workbook(WORKBOOK_PATH)
    if sheet not in wb.sheetnames:
        return jsonify({"error": "Sheet not found"}), 404

    ws = wb[sheet]
    headers, header_row = get_clean_headers(ws)
    header_to_col = {compact_field_name(h): idx for idx, h in enumerate(headers, start=1)}

    for header, value in fields.items():
        col = header_to_col.get(header)
        if col:
            ws.cell(row_number, col).value = value

    wb.save(WORKBOOK_PATH)

    # Rebuild cache immediately after save, but do it in the request so the next view is correct.
    build_cache()

    append_audit("row_saved", {"sheet": sheet, "row_number": row_number, "backup": backup.name})
    return jsonify({"status": "saved", "backup": backup.name})




def first_empty_data_row(ws, header_row):
    for r in range(header_row + 1, ws.max_row + 2):
        vals = [clean_text(ws.cell(r, c).value) for c in range(1, min(ws.max_column, 20) + 1)]
        if not any(vals):
            return r
    return ws.max_row + 1


@app.route("/api/headers/<sheet>")
def api_headers(sheet):
    wb = load_workbook(WORKBOOK_PATH, data_only=True)
    if sheet not in wb.sheetnames:
        return jsonify({"error": "Sheet not found"}), 404

    ws = wb[sheet]
    headers, header_row = get_clean_headers(ws)
    return jsonify({
        "headers": [compact_field_name(h) for h in headers],
        "header_row": header_row
    })


@app.route("/api/add", methods=["POST"])
@role_required("write")
def add_record():
    payload = request.get_json(force=True)
    sheet = payload.get("sheet")
    fields = payload.get("fields", {})

    if not sheet or sheet == "ALL":
        return jsonify({"error": "Please choose a specific sheet first"}), 400

    backup = backup_workbook()

    wb = load_workbook(WORKBOOK_PATH)
    if sheet not in wb.sheetnames:
        return jsonify({"error": "Sheet not found"}), 404

    ws = wb[sheet]
    headers, header_row = get_clean_headers(ws)
    compact_headers = [compact_field_name(h) for h in headers]
    header_to_col = {h: idx for idx, h in enumerate(compact_headers, start=1)}

    new_row = first_empty_data_row(ws, header_row)

    for header, value in fields.items():
        col = header_to_col.get(header)
        if col:
            ws.cell(new_row, col).value = value

    wb.save(WORKBOOK_PATH)

    build_cache()

    append_audit("row_added", {"sheet": sheet, "row_number": new_row, "backup": backup.name})
    return jsonify({
        "status": "added",
        "sheet": sheet,
        "row_number": new_row,
        "backup": backup.name
    })




@app.route("/api/record/<sheet>/<int:row_number>", methods=["DELETE"])
@role_required("write")
def delete_record(sheet, row_number):
    if not sheet or sheet == "ALL":
        return jsonify({"error": "Invalid sheet"}), 400

    backup = backup_workbook()

    wb = load_workbook(WORKBOOK_PATH)
    if sheet not in wb.sheetnames:
        return jsonify({"error": "Sheet not found"}), 404

    ws = wb[sheet]

    if row_number < 1 or row_number > ws.max_row:
        return jsonify({"error": "Invalid row number"}), 400

    ws.delete_rows(row_number, 1)
    wb.save(WORKBOOK_PATH)

    build_cache()

    append_audit("row_deleted", {"sheet": sheet, "row_number": row_number, "backup": backup.name})
    return jsonify({
        "status": "deleted",
        "sheet": sheet,
        "row_number": row_number,
        "backup": backup.name
    })




def extract_unit_label(text):
    text = clean_text(text)
    if not text:
        return "Unknown"

    # Prefer anything after Unit:
    m = re.search(r'Unit\s*:\s*([^|/]+)', text, re.IGNORECASE)
    if m and clean_text(m.group(1)):
        return clean_text(m.group(1))

    # If no Unit: exists, remove Area/Location and use a short fallback.
    text2 = re.sub(r'(?:^|\|)\s*[^|]*(?:Area|Loc\.?|Location)\s*:\s*[^|]*', '', text, flags=re.IGNORECASE)
    text2 = text2.strip(" |/")
    m = re.search(r':\s*([^|/]+)', text2)
    if m and clean_text(m.group(1)):
        return clean_text(m.group(1))

    return text2 or text[:40]


@app.route("/api/unit-graph")
@login_required
def api_unit_graph():
    ensure_cache_thread()

    with CACHE_LOCK:
        if not CACHE_READY:
            return jsonify({
                "ready": False,
                "building": CACHE_BUILDING,
                "error": CACHE_ERROR,
                "nodes": [],
                "links": []
            }), 202

        query = clean_text(request.args.get("q", "")).lower()
        sheet = request.args.get("sheet", "ALL")
        terms = [t for t in re.split(r"\s+", query) if t]

        nodes = {}
        links = []

        def add_node(kind, details):
            label = extract_unit_label(details)
            node_id = f"{kind}:{label.lower()}"
            if node_id not in nodes:
                nodes[node_id] = {
                    "id": node_id,
                    "kind": kind,
                    "label": label,
                    "details": [],
                    "connections": 0
                }
            nodes[node_id]["connections"] += 1
            if details and details not in nodes[node_id]["details"]:
                nodes[node_id]["details"].append(details)
            return node_id

        count = 0
        for record in CACHE_RECORDS:
            if sheet != "ALL" and record["sheet"] != sheet:
                continue

            haystack = record.get("search", "")
            if terms and not all(term in haystack for term in terms):
                continue

            src_details = record.get("source") or record.get("source_simple") or "Unknown source"
            dst_details = record.get("destination") or record.get("destination_simple") or "Unknown destination"

            sid = add_node("source", src_details)
            did = add_node("destination", dst_details)

            links.append({
                "id": f"link-{count}",
                "source": sid,
                "target": did,
                "source_label": nodes[sid]["label"],
                "target_label": nodes[did]["label"],
                "source_details": src_details,
                "destination_details": dst_details,
                "cable_id": record.get("cable_id", ""),
                "sheet": record.get("sheet", ""),
                "row_number": record.get("row_number", ""),
                "network": record.get("network", ""),
                "notes": record.get("notes", "")
            })

            count += 1
            if count >= 500:
                break

        for node in nodes.values():
            node["details"] = node["details"][:10]

        return jsonify({
            "ready": True,
            "nodes": list(nodes.values()),
            "links": links,
            "count": len(links),
            "limited": count >= 500
        })




AUDIT_LOG_PATH = APP_DIR / "audit_log.jsonl"
DEVICE_DB_PATH = APP_DIR / "devices.json"
NMOS_CONFIG_PATH = APP_DIR / "nmos_config.json"


def append_audit(action, details):
    entry = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "action": action,
        "details": details
    }
    with open(AUDIT_LOG_PATH, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry, ensure_ascii=False) + "\n")


def read_audit(limit=200):
    if not AUDIT_LOG_PATH.exists():
        return []
    lines = AUDIT_LOG_PATH.read_text(encoding="utf-8").splitlines()
    entries = []
    for line in lines[-limit:]:
        try:
            entries.append(json.loads(line))
        except Exception:
            pass
    return list(reversed(entries))


def load_device_db():
    if not DEVICE_DB_PATH.exists():
        return []
    try:
        return json.loads(DEVICE_DB_PATH.read_text(encoding="utf-8"))
    except Exception:
        return []


def save_device_db(devices):
    DEVICE_DB_PATH.write_text(json.dumps(devices, indent=2, ensure_ascii=False), encoding="utf-8")


def extract_ips_from_text(text):
    text = clean_text(text)
    ips = []
    # IPv4 only for now because most 2110 sheets use IPv4 multicast/unicast.
    for match in re.findall(r'\b(?:\d{1,3}\.){3}\d{1,3}\b', text):
        try:
            ipaddress.ip_address(match)
            ips.append(match)
        except Exception:
            continue
    return ips


def record_ips(record):
    text = " ".join([
        record.get("network", ""),
        record.get("source", ""),
        record.get("destination", ""),
        record.get("summary", ""),
        " ".join((record.get("fields") or {}).values())
    ])
    return sorted(set(extract_ips_from_text(text)))


def record_vlans(record):
    text = " ".join([
        record.get("network", ""),
        record.get("summary", ""),
        " ".join((record.get("fields") or {}).values())
    ])
    vlans = []
    for m in re.findall(r'\b(?:VLAN|vlan)?\s*([1-9][0-9]{1,3})\b', text):
        try:
            v = int(m)
            if 1 <= v <= 4094:
                vlans.append(str(v))
        except Exception:
            pass
    return sorted(set(vlans))


def infer_flow_type(record):
    text = " ".join([
        record.get("source", ""),
        record.get("destination", ""),
        record.get("network", ""),
        record.get("summary", ""),
        " ".join((record.get("fields") or {}).values())
    ]).lower()

    if any(x in text for x in ["2110-30", "audio", "aes67", "mic", "ifb", "madi"]):
        return "Audio"
    if any(x in text for x in ["2110-40", "anc", "data", "vanc"]):
        return "ANC"
    if any(x in text for x in ["ptp", "grandmaster", "boundary clock"]):
        return "PTP"
    if any(x in text for x in ["nmos", "is-04", "is-05", "registry"]):
        return "NMOS"
    if any(x in text for x in ["2110-20", "video", "cam", "pgm", "preview", "multiview", "mv", "vision"]):
        return "Video"
    if any(x in text for x in ["control", "ember", "snmp", "api"]):
        return "Control"
    return "Unknown"


def filtered_records_from_request():
    ensure_cache_thread()
    with CACHE_LOCK:
        records = list(CACHE_RECORDS)

    query = clean_text(request.args.get("q", "")).lower()
    sheet = request.args.get("sheet", "ALL")
    flow_type = request.args.get("flow_type", "ALL")
    vlan = clean_text(request.args.get("vlan", ""))

    terms = [t for t in re.split(r"\s+", query) if t]
    out = []
    for record in records:
        if sheet != "ALL" and record.get("sheet") != sheet:
            continue
        haystack = record.get("search", "")
        if terms and not all(term in haystack for term in terms):
            continue
        r_type = infer_flow_type(record)
        if flow_type != "ALL" and r_type != flow_type:
            continue
        if vlan and vlan not in record_vlans(record):
            continue
        out.append(record)
    return out


@app.route("/api/validate")
@login_required
def api_validate():
    records = filtered_records_from_request()

    ip_map = {}
    cable_map = {}
    missing = []
    invalid_ips = []

    for record in records:
        cable = clean_text(record.get("cable_id", ""))
        if cable:
            cable_map.setdefault(cable.lower(), []).append(record)

        if not record.get("source") and not record.get("source_simple"):
            missing.append({"type": "Missing source", "sheet": record.get("sheet"), "row": record.get("row_number"), "cable": cable})
        if not record.get("destination") and not record.get("destination_simple"):
            missing.append({"type": "Missing destination", "sheet": record.get("sheet"), "row": record.get("row_number"), "cable": cable})

        for ip in record_ips(record):
            try:
                ipaddress.ip_address(ip)
                ip_map.setdefault(ip, []).append(record)
            except Exception:
                invalid_ips.append({"ip": ip, "sheet": record.get("sheet"), "row": record.get("row_number"), "cable": cable})

    duplicate_ips = []
    for ip, rows in ip_map.items():
        if len(rows) > 1:
            duplicate_ips.append({
                "ip": ip,
                "count": len(rows),
                "items": [{"sheet": r.get("sheet"), "row": r.get("row_number"), "cable": r.get("cable_id")} for r in rows[:25]]
            })

    duplicate_cables = []
    for cable, rows in cable_map.items():
        if len(rows) > 1:
            duplicate_cables.append({
                "cable": rows[0].get("cable_id"),
                "count": len(rows),
                "items": [{"sheet": r.get("sheet"), "row": r.get("row_number")} for r in rows[:25]]
            })

    return jsonify({
        "records_checked": len(records),
        "duplicate_ips": duplicate_ips,
        "duplicate_cables": duplicate_cables,
        "missing": missing[:200],
        "invalid_ips": invalid_ips[:200],
        "summary": {
            "duplicate_ip_count": len(duplicate_ips),
            "duplicate_cable_count": len(duplicate_cables),
            "missing_count": len(missing),
            "invalid_ip_count": len(invalid_ips)
        }
    })


@app.route("/api/flow-summary")
@login_required
def api_flow_summary():
    records = filtered_records_from_request()
    summary = {}
    vlan_summary = {}
    for r in records:
        ft = infer_flow_type(r)
        summary[ft] = summary.get(ft, 0) + 1
        for v in record_vlans(r):
            vlan_summary[v] = vlan_summary.get(v, 0) + 1
    return jsonify({"flow_types": summary, "vlans": vlan_summary})


@app.route("/api/devices", methods=["GET", "POST"])
@role_required("write")
def api_devices():
    if request.method == "GET":
        return jsonify({"devices": load_device_db()})

    payload = request.get_json(force=True)
    devices = load_device_db()
    device = {
        "id": payload.get("id") or f"dev-{int(time.time()*1000)}",
        "name": clean_text(payload.get("name", "")),
        "ip": clean_text(payload.get("ip", "")),
        "vendor": clean_text(payload.get("vendor", "")),
        "role": clean_text(payload.get("role", "")),
        "site": clean_text(payload.get("site", "")),
        "ptp_domain": clean_text(payload.get("ptp_domain", "")),
        "primary_network": clean_text(payload.get("primary_network", "")),
        "secondary_network": clean_text(payload.get("secondary_network", "")),
        "notes": clean_text(payload.get("notes", ""))
    }
    devices = [d for d in devices if d.get("id") != device["id"]]
    devices.append(device)
    save_device_db(devices)
    append_audit("device_saved", device)
    return jsonify({"status": "saved", "device": device})


@app.route("/api/devices/<device_id>", methods=["DELETE"])
@role_required("write")
def api_delete_device(device_id):
    devices = load_device_db()
    before = len(devices)
    devices = [d for d in devices if d.get("id") != device_id]
    save_device_db(devices)
    append_audit("device_deleted", {"id": device_id})
    return jsonify({"status": "deleted", "removed": before - len(devices)})


def ping_host(ip):
    system = platform.system().lower()
    if "windows" in system:
        cmd = ["ping", "-n", "1", "-w", "700", ip]
    else:
        cmd = ["ping", "-c", "1", "-W", "1", ip]
    try:
        result = subprocess.run(cmd, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=2)
        return result.returncode == 0
    except Exception:
        return False


@app.route("/api/status-check")
@role_required("write")
def api_status_check():
    devices = load_device_db()
    results = []
    for d in devices:
        ip = clean_text(d.get("ip", ""))
        if not ip:
            status = "no-ip"
        else:
            status = "online" if ping_host(ip) else "offline"
        item = dict(d)
        item["status"] = status
        results.append(item)
    return jsonify({"devices": results})


@app.route("/api/audit")
@login_required
def api_audit():
    limit = int(request.args.get("limit", 200))
    return jsonify({"entries": read_audit(limit)})


@app.route("/api/nmos-config", methods=["GET", "POST"])
@role_required("write")
def api_nmos_config():
    if request.method == "GET":
        if NMOS_CONFIG_PATH.exists():
            try:
                return jsonify(json.loads(NMOS_CONFIG_PATH.read_text(encoding="utf-8")))
            except Exception:
                pass
        return jsonify({"registry_url": "", "enabled": False, "notes": "Configure your IS-04 registry URL here."})

    payload = request.get_json(force=True)
    config = {
        "registry_url": clean_text(payload.get("registry_url", "")),
        "enabled": bool(payload.get("enabled", False)),
        "notes": clean_text(payload.get("notes", ""))
    }
    NMOS_CONFIG_PATH.write_text(json.dumps(config, indent=2), encoding="utf-8")
    append_audit("nmos_config_saved", config)
    return jsonify({"status": "saved", "config": config})


@app.route("/api/nmos-preview")
@login_required
def api_nmos_preview():
    # Placeholder for live NMOS discovery. This stays safe until a registry URL is configured.
    if NMOS_CONFIG_PATH.exists():
        try:
            config = json.loads(NMOS_CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            config = {}
    else:
        config = {}
    return jsonify({
        "status": "not_connected",
        "message": "NMOS live discovery is prepared but not enabled yet. Configure the IS-04 registry URL first.",
        "config": config,
        "next_steps": [
            "Add IS-04 registry URL",
            "Fetch /x-nmos/query/v1.3/nodes",
            "Fetch devices/sources/flows/senders/receivers",
            "Compare live NMOS inventory against spreadsheet"
        ]
    })


@app.route("/download")
@login_required
def download():
    return send_file(WORKBOOK_PATH, as_attachment=True)


if __name__ == "__main__":
    ensure_cache_thread()
    app.run(host="0.0.0.0", port=5005, debug=True)